import openpyxl

# criar uma planilha(book)
book = openpyxl.Workbook()

# como visualizar páginas existentes
print(book.sheetnames)

# como criar uma pagina
book.create_sheet('Frutas')

# como selecionar uma pagina
frutas_page = book['Frutas']

# adicionar dados
frutas_page.append(['Fruta', 'Quantidade', 'Preço'])
frutas_page.append(['banana', '', 'R$5,00'])
frutas_page.append(['banana', '5', 'R$5,00'])
frutas_page.append(['maçã', '2', 'R$6,00'])
frutas_page.append(['mamão', '1', 'R$4,00'])
frutas_page.append(['limão', '10', 'R$12,00'])
frutas_page.append(['laranja', '15', 'R$35,00'])

# salvar planilha
book.save('planilha de compras.xlsx')
